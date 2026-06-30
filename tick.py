import json
from openpyxl import load_workbook

EXCEL_PATH = "input.xlsx"
JSON_PATH = "existing.json"
OUTPUT_PATH = "output.json"

SHEET_NAME = "Sheet1"

HEADER_ROW = 1
START_ROW = 2

KEY_COL = "C"

# 仕切りで分かれている列ブロック
COLUMN_BLOCKS = [
    ["H", "I", "J", "K"],
    ["N", "O", "P", "Q"],

    # あと2個をここに追加
    # 例:
    # ["T", "U", "V", "W"],
    # ["Z", "AA", "AB", "AC"],
]

with open(JSON_PATH, "r", encoding="utf-8") as f:
    data = json.load(f)

wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb[SHEET_NAME]

def get_cell(col, row):
    return ws[f"{col}{row}"].value

# ブロックを1本の列リストにする
target_cols = []
for block in COLUMN_BLOCKS:
    target_cols.extend(block)

# 各列の見出しを取得
headers = {}
for col in target_cols:
    headers[col] = get_cell(col, HEADER_ROW) or col

for row in range(START_ROW, ws.max_row + 1):
    key_name = get_cell(KEY_COL, row)

    if not key_name:
        continue

    key_name = str(key_name)

    # 既存JSONにそのキーがなければ作る
    if key_name not in data:
        data[key_name] = {}

    for col in target_cols:
        field_name = str(headers[col])
        value = get_cell(col, row)

        data[key_name][field_name] = value

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
